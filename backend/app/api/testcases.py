from typing import List, Optional
from io import BytesIO
import re
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.models.user import User
from app.models.project import Project, ProjectMember
from app.models.requirement import Requirement
from app.models.testcase import TestCase, TestCaseCategory, TestCaseHistory, TestCaseType, TestCaseStatus, TestCasePriority
from sqlalchemy import desc
from app.schemas.testcase import (
    TestCaseCreate, TestCaseUpdate, TestCaseResponse, TestCaseListResponse,
    CategoryCreate, CategoryUpdate, CategoryResponse,
    TestCaseBatchDeleteRequest
)
from app.utils.dependencies import get_current_user

# 字段映射常量
EXPORT_COLUMNS = [
    '用例目录', '模块', '功能', '用例名称', '前置条件', '用例步骤', 
    '测试数据', '预期结果', '实际结果', '用例类型', '用例状态', '用例等级', '需求ID'
]

TYPE_MAP = {
    '功能测试': TestCaseType.FUNCTIONAL,
    '性能测试': TestCaseType.PERFORMANCE,
    '安全测试': TestCaseType.SECURITY,
    '兼容性测试': TestCaseType.COMPATIBILITY,
    '冒烟测试': TestCaseType.SMOKE,
    '回归测试': TestCaseType.REGRESSION,
}
TYPE_MAP_REVERSE = {v: k for k, v in TYPE_MAP.items()}

STATUS_MAP = {
    '通过': TestCaseStatus.PASSED,
    '已通过': TestCaseStatus.PASSED,
    '不通过': TestCaseStatus.FAILED,
    '已失败': TestCaseStatus.FAILED,
    '未执行': TestCaseStatus.NOT_EXECUTED,
    '待测试': TestCaseStatus.NOT_EXECUTED,
}
STATUS_MAP_REVERSE = {
    TestCaseStatus.PASSED: '通过',
    TestCaseStatus.FAILED: '不通过',
    TestCaseStatus.NOT_EXECUTED: '未执行',
}

PRIORITY_MAP = {
    '高': TestCasePriority.HIGH,
    '中': TestCasePriority.MEDIUM,
    '低': TestCasePriority.LOW,
}
PRIORITY_MAP_REVERSE = {v: k for k, v in PRIORITY_MAP.items()}


def strip_html_tags(html: str) -> str:
    """去除 HTML 标签，保留纯文本内容"""
    if not html:
        return ''
    # 将 <br>, <br/>, </p>, </div>, </li> 等标签替换为换行
    text = re.sub(r'<br\s*/?>', '\n', html, flags=re.IGNORECASE)
    text = re.sub(r'</(?:p|div|li|h[1-6])>', '\n', text, flags=re.IGNORECASE)
    # 移除所有其他 HTML 标签
    text = re.sub(r'<[^>]+>', '', text)
    # 解码 HTML 实体
    text = text.replace('&nbsp;', ' ').replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&').replace('&quot;', '"')
    # 清理多余的空行
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


router = APIRouter(prefix="/api/testcases", tags=["testcases"])


# ========== TestCase Endpoints ==========

@router.get("/", response_model=TestCaseListResponse)
def get_testcases(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    project_id: Optional[int] = None,
    category_id: Optional[int] = None,
    requirement_id: Optional[int] = None,
    sprint_id: Optional[int] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    creator_id: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get testcases with filters and pagination"""
    query = db.query(TestCase)
    
    # Filter by project access
    accessible_projects = db.query(ProjectMember.project_id).filter(
        ProjectMember.user_id == current_user.id
    ).all()
    accessible_project_ids = [p[0] for p in accessible_projects]
    
    if accessible_project_ids:
        query = query.filter(TestCase.project_id.in_(accessible_project_ids))
    
    # Apply filters
    if project_id:
        query = query.filter(TestCase.project_id == project_id)
    if category_id:
        # 获取该目录及其所有子目录的 ID
        def get_category_ids_recursive(cat_id: int) -> List[int]:
            """递归获取目录及其所有子目录的 ID"""
            ids = [cat_id]
            children = db.query(TestCaseCategory.id).filter(
                TestCaseCategory.parent_id == cat_id
            ).all()
            for (child_id,) in children:
                ids.extend(get_category_ids_recursive(child_id))
            return ids
        
        all_category_ids = get_category_ids_recursive(category_id)
        query = query.filter(TestCase.category_id.in_(all_category_ids))
    if requirement_id:
        query = query.filter(TestCase.requirement_id == requirement_id)
    if sprint_id:
        query = query.filter(TestCase.sprint_id == sprint_id)
    if type:
        query = query.filter(TestCase.type == type)
    if status:
        query = query.filter(TestCase.status == status)
    if priority:
        query = query.filter(TestCase.priority == priority)
    if creator_id:
        query = query.filter(TestCase.creator_id == creator_id)
    if search:
        query = query.filter(or_(
            TestCase.name.contains(search),
            TestCase.case_number.contains(search)
        ))
    
    # Get total count
    total = query.count()
    
    # Pagination
    skip = (page - 1) * page_size
    testcases = query.order_by(TestCase.created_at.desc()).offset(skip).limit(page_size).all()
    
    return {
        "items": testcases,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.post("/", response_model=TestCaseResponse, status_code=status.HTTP_201_CREATED)
def create_testcase(
    testcase_data: TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new testcase"""
    # Verify project exists and user has access
    project = db.query(Project).filter(Project.id == testcase_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if user is a member
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == testcase_data.project_id,
        ProjectMember.user_id == current_user.id
    ).first()
    if not member and project.creator_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Create testcase with placeholder number
    testcase = TestCase(
        project_id=testcase_data.project_id,
        category_id=testcase_data.category_id,
        requirement_id=testcase_data.requirement_id,
        sprint_id=testcase_data.sprint_id,
        case_number="TEMP",  # Will be updated after getting ID
        name=testcase_data.name,
        module=testcase_data.module,
        feature=testcase_data.feature,
        type=testcase_data.type,
        status=testcase_data.status,
        priority=testcase_data.priority,
        precondition=testcase_data.precondition,
        steps=testcase_data.steps,
        test_data=testcase_data.test_data,
        expected_result=testcase_data.expected_result,
        actual_result=testcase_data.actual_result,
        creator_id=current_user.id
    )
    
    db.add(testcase)
    db.flush()  # Get the ID
    
    # Update case number using the database ID
    testcase.case_number = f"TC{testcase.id}"
    
    db.commit()
    db.refresh(testcase)
    
    return testcase


@router.post("/batch-delete", status_code=status.HTTP_204_NO_CONTENT)
def batch_delete_testcases(
    data: TestCaseBatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Batch delete testcases"""
    testcases = db.query(TestCase).filter(TestCase.id.in_(data.ids)).all()
    
    for testcase in testcases:
        db.delete(testcase)
    
    db.commit()
    return None


# ========== Category Endpoints ==========
# NOTE: These routes MUST be defined before /{testcase_id} routes to avoid path conflicts

@router.get("/categories", response_model=List[CategoryResponse])
def get_categories(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get testcase categories for a project"""
    categories = db.query(TestCaseCategory).filter(
        TestCaseCategory.project_id == project_id
    ).order_by(TestCaseCategory.order, TestCaseCategory.id).all()
    
    return categories


@router.post("/categories", response_model=CategoryResponse, status_code=status.HTTP_201_CREATED)
def create_category(
    category_data: CategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new category"""
    # Verify project exists
    project = db.query(Project).filter(Project.id == category_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get max order
    max_order = db.query(TestCaseCategory).filter(
        TestCaseCategory.project_id == category_data.project_id,
        TestCaseCategory.parent_id == category_data.parent_id
    ).count()
    
    category = TestCaseCategory(
        project_id=category_data.project_id,
        parent_id=category_data.parent_id,
        name=category_data.name,
        order=max_order
    )
    
    db.add(category)
    db.commit()
    db.refresh(category)
    
    return category


@router.put("/categories/{category_id}", response_model=CategoryResponse)
def update_category(
    category_id: int,
    category_data: CategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update category"""
    category = db.query(TestCaseCategory).filter(TestCaseCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    update_data = category_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(category, field, value)
    
    db.commit()
    db.refresh(category)
    
    return category


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete category (testcases in this category will become uncategorized)"""
    category = db.query(TestCaseCategory).filter(TestCaseCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Move testcases to uncategorized
    db.query(TestCase).filter(TestCase.category_id == category_id).update(
        {TestCase.category_id: None}
    )
    
    # Move child categories to parent
    db.query(TestCaseCategory).filter(TestCaseCategory.parent_id == category_id).update(
        {TestCaseCategory.parent_id: category.parent_id}
    )
    
    db.delete(category)
    db.commit()
    return None


# ========== Import/Export Endpoints ==========
# NOTE: These routes MUST be defined before /{testcase_id} routes to avoid path conflicts

@router.get("/template")
def download_template():
    """下载测试用例导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例导入模板"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入示例数据
    sample_data = [
        '用户模块', '登录功能', '登录验证', '登录功能 - 正常流程验证',
        '1、用户已注册\n2、已进入登录页面',
        '1、输入用户名\n2、输入密码\n3、点击登录按钮',
        '用户名: test001, 密码: 123456',
        '1、登录成功\n2、跳转到首页',
        '',  # 实际结果
        '功能测试', '未执行', '高', 'R1'
    ]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
    
    # 调整列宽
    column_widths = [15, 15, 15, 30, 30, 40, 25, 40, 25, 12, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 80
    
    # 返回文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=testcase_template.xlsx"}
    )


@router.get("/export")
def export_testcases(
    project_id: int,
    category_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出测试用例到 Excel"""
    # 查询测试用例
    query = db.query(TestCase).filter(TestCase.project_id == project_id)
    if category_id:
        query = query.filter(TestCase.category_id == category_id)
    testcases = query.order_by(TestCase.created_at.desc()).all()
    
    # 获取目录映射
    categories = db.query(TestCaseCategory).filter(
        TestCaseCategory.project_id == project_id
    ).all()
    category_map = {c.id: c.name for c in categories}
    
    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row, tc in enumerate(testcases, 2):
        # 获取需求编号
        req_number = None
        if tc.requirement_id:
            req = db.query(Requirement).filter(Requirement.id == tc.requirement_id).first()
            if req:
                req_number = req.requirement_number
        
        row_data = [
            category_map.get(tc.category_id, '') if tc.category_id else '',
            tc.module or '',
            tc.feature or '',
            tc.name or '',
            strip_html_tags(tc.precondition or ''),
            strip_html_tags(tc.steps or ''),
            strip_html_tags(tc.test_data or ''),
            strip_html_tags(tc.expected_result or ''),
            strip_html_tags(tc.actual_result or ''),
            TYPE_MAP_REVERSE.get(tc.type, str(tc.type.value) if tc.type else ''),
            STATUS_MAP_REVERSE.get(tc.status, str(tc.status.value) if tc.status else ''),
            PRIORITY_MAP_REVERSE.get(tc.priority, str(tc.priority.value) if tc.priority else ''),
            req_number or '',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
    
    # 调整列宽
    column_widths = [15, 15, 15, 30, 30, 40, 25, 40, 25, 12, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # 返回文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=testcases_export.xlsx"}
    )


@router.post("/import")
async def import_testcases(
    file: UploadFile = File(...),
    project_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """从 Excel 导入测试用例"""
    # 验证文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 或 .xls 格式的文件")
    
    # 验证空间存在
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="空间不存在")
    
    # 检查用户权限
    member = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == current_user.id
    ).first()
    if not member and project.creator_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权访问该空间")
    
    # 读取 Excel
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件读取失败: {str(e)}")
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 获取现有目录映射
    categories = db.query(TestCaseCategory).filter(
        TestCaseCategory.project_id == project_id
    ).all()
    category_name_map = {c.name: c.id for c in categories}
    
    # 获取需求映射
    requirements = db.query(Requirement).filter(
        Requirement.project_id == project_id
    ).all()
    req_number_map = {r.requirement_number: r.id for r in requirements}
    
    # 导入统计
    success_count = 0
    error_count = 0
    errors = []
    
    # 处理每一行数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not any(row):
            continue
        
        # 构建数据字典
        row_data = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = value
        
        # 验证必填字段
        name = row_data.get('用例名称')
        if not name:
            error_count += 1
            errors.append(f"第 {row_idx} 行: 用例名称不能为空")
            continue
        
        try:
            # 处理目录
            category_id = None
            category_name = row_data.get('用例目录')
            if category_name:
                if category_name in category_name_map:
                    category_id = category_name_map[category_name]
                else:
                    # 自动创建目录
                    new_category = TestCaseCategory(
                        project_id=project_id,
                        name=category_name,
                        order=len(category_name_map)
                    )
                    db.add(new_category)
                    db.flush()
                    category_id = new_category.id
                    category_name_map[category_name] = category_id
            
            # 处理需求ID
            requirement_id = None
            req_number = row_data.get('需求ID')
            if req_number:
                requirement_id = req_number_map.get(str(req_number))
            
            # 处理类型、状态、优先级
            tc_type = TYPE_MAP.get(row_data.get('用例类型', ''), TestCaseType.FUNCTIONAL)
            tc_status = STATUS_MAP.get(row_data.get('用例状态', ''), TestCaseStatus.NOT_EXECUTED)
            tc_priority = PRIORITY_MAP.get(row_data.get('用例等级', ''), TestCasePriority.MEDIUM)
            
            # 创建测试用例
            testcase = TestCase(
                project_id=project_id,
                category_id=category_id,
                requirement_id=requirement_id,
                case_number="TEMP",
                name=name,
                module=row_data.get('模块'),
                feature=row_data.get('功能'),
                type=tc_type,
                status=tc_status,
                priority=tc_priority,
                precondition=row_data.get('前置条件'),
                steps=row_data.get('用例步骤'),
                test_data=row_data.get('测试数据'),
                expected_result=row_data.get('预期结果'),
                actual_result=row_data.get('实际结果'),
                creator_id=current_user.id
            )
            db.add(testcase)
            db.flush()
            testcase.case_number = f"TC{testcase.id}"
            success_count += 1
            
        except Exception as e:
            error_count += 1
            errors.append(f"第 {row_idx} 行: {str(e)}")
    
    db.commit()
    
    return {
        "success": True,
        "message": f"导入完成，成功 {success_count} 条，失败 {error_count} 条",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:10] if errors else []  # 最多返回10条错误
    }


# ========== Dynamic TestCase Endpoints ==========
# NOTE: These routes with {testcase_id} MUST come after all static routes

@router.get("/{testcase_id}", response_model=TestCaseResponse)
def get_testcase(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get testcase by ID"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="TestCase not found")
    return testcase


@router.put("/{testcase_id}", response_model=TestCaseResponse)
def update_testcase(
    testcase_id: int,
    testcase_data: TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update testcase"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="TestCase not found")
    
    # Update fields and record history
    update_data = testcase_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        old_value = getattr(testcase, field)
        if old_value != value:
            old_str = old_value.value if hasattr(old_value, 'value') else str(old_value) if old_value is not None else None
            new_str = value.value if hasattr(value, 'value') else str(value) if value is not None else None
            history_entry = TestCaseHistory(
                testcase_id=testcase_id,
                field=field,
                old_value=old_str,
                new_value=new_str,
                changed_by=current_user.id,
            )
            db.add(history_entry)
        setattr(testcase, field, value)
    
    db.commit()
    db.refresh(testcase)
    
    return testcase


@router.get("/{testcase_id}/history")
def get_testcase_history(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get testcase history"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="TestCase not found")

    history = db.query(TestCaseHistory).filter(
        TestCaseHistory.testcase_id == testcase_id
    ).order_by(desc(TestCaseHistory.changed_at)).all()
    
    return [
        {
            "id": h.id,
            "field": h.field,
            "old_value": h.old_value,
            "new_value": h.new_value,
            "changed_at": h.changed_at,
            "user": {"id": h.user.id, "username": h.user.username} if h.user else None
        }
        for h in history
    ]


@router.delete("/{testcase_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_testcase(
    testcase_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete testcase"""
    testcase = db.query(TestCase).filter(TestCase.id == testcase_id).first()
    if not testcase:
        raise HTTPException(status_code=404, detail="TestCase not found")
    
    db.delete(testcase)
    db.commit()
    return None
